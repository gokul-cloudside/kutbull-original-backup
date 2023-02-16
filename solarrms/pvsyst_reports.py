from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, Color
from solarrms.models import SolarPlant, PlantSummaryDetails, PVSystInfo
from django.contrib.auth.models import User
import datetime
import pandas as pd
import numpy as np
from dateutil import parser
from solarrms.api_views import update_tz
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, FormatObject
from openpyxl import formatting, styles
import calendar, pytz
from solarrms.cron_send_reports import send_detailed_report_with_attachment_for_pvsyst_cleanmax
import logging

logger = logging.getLogger('dataglen.rest_views')
logger.setLevel(logging.DEBUG)

def update_timestamps(starttime, endtime, plant):
    try:
        try:
            starttime = starttime.astimezone(pytz.timezone(plant.metadata.dataTimezone)).replace(hour=0, minute=0, second=0,
                                                                                                 microsecond=0)
            endtime = endtime.astimezone(pytz.timezone(plant.metadata.dataTimezone)).replace(hour=0, minute=0, second=0,
                                                                                             microsecond=0)
        except:
            starttime = starttime.replace(tzinfo=pytz.UTC).astimezone(pytz.timezone(plant.metadata.dataTimezone)).replace(
                hour=0, minute=0, second=0, microsecond=0)
            endtime = endtime.replace(tzinfo=pytz.UTC).astimezone(pytz.timezone(plant.metadata.dataTimezone)).replace(
                hour=0, minute=0, second=0, microsecond=0)
        return starttime, endtime
    except:
        return starttime, endtime


# Get regular data from PlantSummaryDetails on daily basis.
def get_plant_details_raw_data(starttime, endtime, plant):
    try:
        starttime, endtime = update_timestamps(starttime, endtime, plant)
        # endtime = endtime.replace(hour=23)
        df0 = pd.DataFrame()
        plant_summary_values = PlantSummaryDetails.objects.filter(timestamp_type='BASED_ON_REQUEST_ARRIVAL',
                                                                  count_time_period=86400,
                                                                  identifier=str(plant.slug),
                                                                  ts__gte=starttime,
                                                                  ts__lte=endtime).order_by('ts').values_list(
            'generation', 'average_irradiation', 'performance_ratio', 'specific_yield', 'ts')
        df = pd.DataFrame(list(plant_summary_values),
                          columns=['generation', 'average_irradiation', 'performance_ratio', 'specific_yield', 'ts'])
        # df['ts'] = df['ts'].apply(lambda x: x.date())
        df = df.rename(columns={'ts': 'Date'})
        # NOTE : REPLACING ZERO WITH NAN. RETHINK
        # df = df.replace(0, np.nan)
        no_of_rows_in_df = df.shape[0]
        df0['Plant'] = [plant.slug] * no_of_rows_in_df
        df0['Location'] = [plant.location] * no_of_rows_in_df
        df0['Capacity (kWp)'] = [plant.capacity] * no_of_rows_in_df
        df0['Commissioning Date'] = [plant.commissioned_date] * no_of_rows_in_df
        df = df0.join(df)
        df = df.sort("Date").reset_index(drop=True)
        return df
    except Exception as exception:
        print "EXCEPTION IN get_plant_details_raw_data", str(exception)
        return pd.DataFrame()

# Send start time and end time. Generates [start day of month, end day of month] list of list.
def get_list_of_days_per_month(st, et):
    try:
        # st, et = update_timestamps(st, et, plants[0])
        total_days = (et - st).days + 1
        day1 = st
        l = []
        flag = 0
        days = 0
        # for days in range(total_days):
        while days < total_days:
            if day1 == st and flag == 0:
                flag = 1
                no_of_days_in_this_month = calendar.monthrange(st.year, st.month)[1]
                if et.month == st.month and et.year == st.year:
                    day2 = et.replace(hour=23,minute=59,second=59)
                    days = total_days
                else:
                    day2 = st.replace(day=no_of_days_in_this_month,hour=23,minute=59,second=59)
                    days = days + (no_of_days_in_this_month - 0 if st.day==1 else st.day)
                print "===============1"
                print day1, day2
                l.append([day1, day2])
            else:
                next_month = day1.month + 1
                next_year = day1.year
                if next_month == 13:
                    next_month = 1
                    next_year = day1.year + 1
                day1 = day1.replace(day=1, month=next_month, year=next_year)
                no_of_days_in_this_month = calendar.monthrange(day1.year, day1.month)[1]
                day2 = day1.replace(day=no_of_days_in_this_month,hour=23,minute=59,second=59)
                if (total_days - days) < no_of_days_in_this_month:
                    day2 = day1.replace(day=(total_days - days)-1,hour=23,minute=59,second=59)
                    days = total_days
                print "===============2"
                print day1, day2
                l.append([day1, day2])
                days = days + no_of_days_in_this_month
        return l
    except Exception as e:
        print "EXCEPTION AS " + str(e)
        return []


def get_pvsyst_data_for_sheets(st, et, plants):
    try:
        # st, et = update_timestamps(st, et, plants[0])
        l = get_list_of_days_per_month(st, et)
        df_dict = {}

        for plant in plants:
            df_appender = pd.DataFrame()
            for days in l:
                df = get_pvsyst_raw_data_in_one_month(days[0], days[1], plant)
                df_appender = df_appender.append(df)
            df_dict[plant.slug] = df_appender
        return df_dict
    except Exception as e:
        print 'EXCEPTION' + str(e)
        return {}


def get_pvsyst_raw_data_in_one_month(st, et, plant):
    try:
        df_combined_result_single_plant = pd.DataFrame()
        if st.month == et.month and st.year == et.year:
            pv_sys_info_generation = PVSystInfo.objects.filter(plant=plant,
                                                               parameterName__in=['PRODUCED_ENERGY',
                                                                                  'GHI_IRRADIANCE',
                                                                                  'PERFORMANCE_RATIO'],
                                                               timePeriodType='MONTH',
                                                               timePeriodDay=0,
                                                               timePeriodValue=st.month,
                                                               timePeriodYear__in=[st.year, 0]).values_list(
                'timePeriodValue', 'parameterName', 'parameterValue', 'solar_group_id')

            d = {}
            if len(pv_sys_info_generation) > 0:
                df = pd.DataFrame(list(pv_sys_info_generation))
                df.columns = ['timePeriodValue', 'parameterName', 'parameterValue', 'solar_group_id']
                # As there can be different groups we need to combine data for multiple groups of a plant
                gdf = df.groupby(['parameterName', 'timePeriodValue'])
                d['Plant'] = plant.slug
                for g in gdf:
                    if g[0][0] in ['PERFORMANCE_RATIO','GHI_IRRADIANCE']:
                        d[g[0][0]] = g[1].mean()['parameterValue']
                    else:
                        d[g[0][0]] = g[1].sum()['parameterValue']
            else:
                print "No data in PVSystInfo", plant
            if len(d.keys()) > 0:
                no_of_days_of_month_in_st = (st.replace(month=(st.month + 1)) - st).days
                # days_in_st_month = no_of_days_of_month_in_st - st.day + 1
                produced_energy_per_day = (d['PRODUCED_ENERGY'] / no_of_days_of_month_in_st) * 1000
                ghi_irradiance_per_day = d['GHI_IRRADIANCE'] / no_of_days_of_month_in_st
                # if plant.slug == 'omya':
                #     ghi_irradiance_per_day = ghi_irradiance_per_day / 7.0
                performance_ratio_per_day = d['PERFORMANCE_RATIO']

                list_of_data = []
                for day in range(st.day, et.day + 1):
                    list_of_data.append(
                        [st.replace(day=day), plant.slug, produced_energy_per_day, ghi_irradiance_per_day,
                         performance_ratio_per_day])

                df_combined_result_single_plant = pd.DataFrame(list_of_data,
                                                               columns=['Date', 'Plant', 'produced_energy_per_day',
                                                                        'ghi_irradiance_per_day',
                                                                        'performance_ratio_per_day'])

        else:
            print "This function accepts only single month values. Start Time and End Time should belong to same month"
        return df_combined_result_single_plant
    except Exception as e:
        print "Exception in get_pvsyst_data : " + str(e)
        return pd.DataFrame()


def calculate_aged_generation(plants):
    try:
        x = {}
        for plant in plants:
            com_date = plant.commissioned_date
            if type(plant.commissioned_date) is datetime.date:
                today = datetime.datetime.now().date()
                tdelta = today - com_date
                plant_age = tdelta.days / 365
                expected_gen = 100
                expected_gen_reduced_as_aging = 0
                if plant_age == 0:
                    # print "less than one year old"
                    expected_gen_reduced_as_aging = expected_gen * 0.975  # reduced 2.5% for first year
                else:
                    expected_gen_reduced_as_aging = expected_gen * 0.975  # reduced 2.5% for first year
                    for i in range(plant_age):
                        # print expected_gen_reduced_as_aging
                        expected_gen_reduced_as_aging = expected_gen_reduced_as_aging * 0.993  # reduced 0.7% after second year
                x[plant.slug] = (expected_gen_reduced_as_aging)
            else:
                # print "in else"
                x[plant.slug] = 0
        return x
    except Exception as e:
        print "Exception in calculate_aged_generation :" + str(e)
        return {}


# Main Function
def df_for_pvsyst_analysis(st, et, user, plants):
    try:

        df_dict_meta = {}
        commissioned_date_dict = {}
        no_of_rows = (et - st).days + 1

        # Add data from Plant Summary Details ['Plant', 'Location', 'Capacity (kWp)', 'Date' ,'Actual Total Gen','Actual Irradiation', 'Actual PR','Specific Yield (Per kWp gen)']
        for plant in plants:
            st, et = update_timestamps(st, et, plant)
            commissioned_date_dict[plant.slug] = plant.commissioned_date
            df_plant_meta = get_plant_details_raw_data(st, et, plant)
            df_plant_meta['Date'] = df_plant_meta['Date'].map(lambda x: (
                x.replace(tzinfo=pytz.UTC).astimezone(pytz.timezone(plant.metadata.dataTimezone)).replace(
                    hour=0, minute=0, second=0, microsecond=0)).date())
            if df_plant_meta.empty:
                print "No Data Nothing can be done. There is no data for the plant : ", plant
            else:
                df_dict_meta[plant.slug] = df_plant_meta

        print "one tzinfo done for df_plant_meta get platn details raw data"
        # Add data from table PVSystInfo ['Date','Plant', 'produced_energy_per_day', 'ghi_irradiance_per_day','performance_ratio_per_day']
        df_pvsyst_data_dict = get_pvsyst_data_for_sheets(st, et, plants)
        print "df pvsyst ",df_pvsyst_data_dict
        plants_with_pvsyst = []
        for key in df_dict_meta:
            if df_pvsyst_data_dict[key].empty:
                print "Do nothing", key
            else:
                df_pvsyst_data_dict[key]['Date'] = df_pvsyst_data_dict[key]['Date'].map(lambda x: (
                    x.replace(tzinfo=pytz.UTC).astimezone(pytz.timezone(plant.metadata.dataTimezone)).replace(
                        hour=0, minute=0, second=0, microsecond=0)).date())
                df_dict_meta[key] = df_dict_meta[key].merge(df_pvsyst_data_dict[key], on=['Date', 'Plant'], how='outer')
                plants_with_pvsyst.append(key)

        df_dict_age = {}
        print 'playing with dates FOR AGED GENERATION'
        # This works only if Commissioning  Date is entered in database
        dict_age = calculate_aged_generation(plants)
        for slug in plants_with_pvsyst:
            df_1 = pd.DataFrame()
            df_1['Date'] = [(st + datetime.timedelta(days=day)).date() for day in range((et - st).days + 1)]
            df_1['Expected gen after aging factor'] = [dict_age[slug]] * no_of_rows
            df_dict_age[slug] = df_1
        for key in plants_with_pvsyst:
            df_dict_meta[key] = df_dict_meta[key].merge(df_dict_age[key], on='Date', how='outer')

        # Some Mathematical manipulations
        for key in df_dict_meta:
            if 'produced_energy_per_day' in list(df_dict_meta[key]):
                df_dict_meta[key]['Expected gen after aging factor'] = df_dict_meta[key][
                                                                           'Expected gen after aging factor'] * \
                                                                       df_dict_meta[key][
                                                                           'produced_energy_per_day'] / 100
            if 'performance_ratio_per_day' in list(df_dict_meta[key]):
                df_dict_meta[key]['performance_ratio_per_day'] = df_dict_meta[key]['performance_ratio_per_day'] * 100
            if 'performance_ratio' in list(df_dict_meta[key]):
                df_dict_meta[key]['performance_ratio'] = df_dict_meta[key]['performance_ratio'] * 100
            else:
                print "Do nothing"

        # Renaming and Some more Mathematical Manipulations
        for key in df_dict_meta:
            print 'RENAMING OPERATIONS FOR KEY ========', key
            df_dict_meta[key] = df_dict_meta[key].rename(
                columns={'generation': 'Actual Total Gen', 'average_irradiation': 'Actual Irradiation',
                         'performance_ratio': 'Actual PR', 'specific_yield': 'Specific Yield (Per kWp gen)',
                         'produced_energy_per_day': 'Expected Gen (As per PV syst)',
                         'ghi_irradiance_per_day': 'Expected Irradiation',
                         'performance_ratio_per_day': 'Expected PR (As per PV syst)',
                         'Expected gen after aging factor': 'Expected generation considering 2.5% for 1st yr. & 0.7% yearly  after that.'})

            if 'Expected Irradiation' in list(df_dict_meta[key]):
                print "for key calculating expected gen :",key
                df_dict_meta[key]['Expected Gen corrected to actual irradiation'] = \
                    df_dict_meta[key]['Expected generation considering 2.5% for 1st yr. & 0.7% yearly  after that.'] / \
                    (df_dict_meta[key]['Expected Irradiation'] / df_dict_meta[key]['Actual Irradiation'])
            else:
                print "Do nothing"

        print "REARRANGING SEQUENCE OF COLUMNS=========>"
        restructure = [u'Plant', u'Location', u'Capacity (kWp)', u'Date', u'Actual Total Gen', u'Actual Irradiation',
                       u'Actual PR', u'Specific Yield (Per kWp gen)',
                       u'Expected Gen (As per PV syst)',
                       u'Expected generation considering 2.5% for 1st yr. & 0.7% yearly  after that.',
                       u'Expected Irradiation', u'Expected Gen corrected to actual irradiation',
                       u'Expected PR (As per PV syst)']
        restructure2 = [u'Plant', u'Location', u'Capacity (kWp)', u'Date', u'Actual Total Gen', u'Actual Irradiation',
                        u'Actual PR', u'Specific Yield (Per kWp gen)']
        for key in df_dict_meta:
            if 'Expected Gen (As per PV syst)' in list(df_dict_meta[key]):
                df_dict_meta[key] = df_dict_meta[key][restructure]
                df_dict_meta[key]['Specific Yield (Per kWp gen) (Expected)'] = df_dict_meta[key][
                                                                                   'Expected generation considering 2.5% for 1st yr. & 0.7% yearly  after that.'] / (
                                                                               df_dict_meta[key]['Capacity (kWp)'])
                df_dict_meta[key]['% Achievement (irradiance)'] = df_dict_meta[key]['Actual Irradiation'] / \
                                                                  df_dict_meta[key]['Expected Irradiation'] * 100
                df_dict_meta[key]['% of Achievement in gen with respect to corrected Irradiance'] = df_dict_meta[key][
                                                                                                        'Actual Total Gen'] / \
                                                                                                    df_dict_meta[key][
                                                                                                        'Expected Gen corrected to actual irradiation'] * 100
                df_dict_meta[key]['% of Achievement (gen)'] = df_dict_meta[key]['Actual Total Gen'] / df_dict_meta[key][
                    'Expected generation considering 2.5% for 1st yr. & 0.7% yearly  after that.'] * 100
                df_dict_meta[key]['% of Achievement (PR)'] = df_dict_meta[key]['Actual PR'] / df_dict_meta[key][
                    'Expected PR (As per PV syst)'] * 100
            else:
                df_dict_meta[key] = df_dict_meta[key][restructure2]
            cols_filtered = list(
                set(list(df_dict_meta[key])) - set(['Date', 'Plant', 'Location', 'Capacity (kWp)', 'Date', ]))

            for col in cols_filtered:
                df_dict_meta[key][col] = df_dict_meta[key][col].apply(lambda x: round(x, 2))

        # =============== CALCULATION FOR SUMMARY SHEET ==================
        dict_summary = {}
        fix_list = ['Plant', 'Location', 'Capacity (kWp)', 'Commissioning Date']
        normal_data_list = ['Actual Total Gen',
                            'Actual Irradiation',
                            'Actual PR',
                            'Specific Yield (Per kWp gen)']
        pvsyst_data_list = ['Expected Gen (As per PV syst)',
                            'Expected generation considering 2.5% for 1st yr. & 0.7% yearly  after that.',
                            'Expected Irradiation',
                            'Expected Gen corrected to actual irradiation',
                            'Expected PR (As per PV syst)',
                            'Specific Yield (Per kWp gen) (Expected)',
                            '% Achievement (irradiance)',
                            '% of Achievement in gen with respect to corrected Irradiance',
                            '% of Achievement (gen)',
                            '% of Achievement (PR)']
        list_of_columns = fix_list + normal_data_list + pvsyst_data_list
        sumlist1 = ['Actual Total Gen']
        meanlist1 = ['Actual PR','Actual Irradiation','Specific Yield (Per kWp gen)']
        sumlist2 = ['Expected Gen (As per PV syst)',
                    'Expected generation considering 2.5% for 1st yr. & 0.7% yearly  after that.',
                    'Expected Gen corrected to actual irradiation']
        meanlist2 = ['Expected PR (As per PV syst)','Expected Irradiation','Specific Yield (Per kWp gen) (Expected)',
                     '% Achievement (irradiance)',
                     '% of Achievement in gen with respect to corrected Irradiance',
                     '% of Achievement (gen)',
                     '% of Achievement (PR)']

        # creating dictionary with colums names as key and empty list as value
        for col in list_of_columns:
            dict_summary[col] = []

        # adding values to all dictionary keys. basically creating columns for dataframe 'Summary'
        for key in df_dict_meta:
            for col in fix_list:
                if col == 'Commissioning Date':
                    dict_summary[col].append(commissioned_date_dict[key])
                else:
                    dict_summary[col].append(df_dict_meta[key][col].iloc[0])

            # Check if plant summary data is there or not
            if 'Actual Total Gen' in list(df_dict_meta[key]):
                for cols in normal_data_list:
                    if cols in sumlist1:
                        dict_summary[cols].append(df_dict_meta[key][cols].sum())
                    elif cols in meanlist1:
                        dict_summary[cols].append(df_dict_meta[key][cols].mean())
            else:
                for cols in normal_data_list:
                    dict_summary[cols].append('')

            # Check if PVSyst data is there or not
            if 'Expected Gen (As per PV syst)' in list(df_dict_meta[key]):
                for cols in pvsyst_data_list:
                    if cols in sumlist2:
                        dict_summary[cols].append(df_dict_meta[key][cols].sum())
                    elif cols in meanlist2:
                        dict_summary[cols].append(df_dict_meta[key][cols].mean())
            else:
                for cols in pvsyst_data_list:
                    dict_summary[cols].append('')

        df_dict_meta['Summary'] = pd.DataFrame(dict_summary)
        # Reorganizing column ordering in data frame
        df_dict_meta['Summary'] = df_dict_meta['Summary'][list_of_columns]
        print "=============Plants slug to Plant name=============="
        for key in df_dict_meta:
            df_dict_meta[key]['Plant'] = df_dict_meta[key]['Plant'].apply(lambda x: (plants.get(slug=x)).name)
        return df_dict_meta
    except Exception as e:
        print "Exception in df_for_pvsyst_analysis : " + str(e)
        return pd.DataFrame()

# For Openpyxl excel formatting
def merge_and_style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """
    try:
        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        if alignment:
            ws.merge_cells(cell_range)
            first_cell.alignment = alignment

        rows = ws[cell_range]
        if font:
            first_cell.font = font

        for cell in rows[0]:
            cell.border = cell.border + top
        for cell in rows[-1]:
            cell.border = cell.border + bottom

        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = l.border + left
            r.border = r.border + right
            if fill:
                for c in row:
                    c.fill = fill
    except Exception as e:
        print "EXCEPTION IN merge_and_style_range "+str(e)

# Create Customised Excel For Pysyst Achievement Analysis
def ExcelPvsystFormat(df, ws, days_range, no_of_days):
    try:
        df_c = df.columns
        # Create empty first row for merging and headlines
        first_row = ['' for col in range(df.shape[1])]
        # insert this first row in dataframe
        df.columns = first_row
        df.loc[-1] = df_c
        df.index = df.index + 1  # shifting index
        df = df.sort_index()  # sorting by index

        # adding dataframe to worksheet, row by row
        for r in dataframe_to_rows(df, index=True, header=True):
            ws.append(r)

        # All Styles Defined :-
        blueFill = PatternFill(start_color='8FAADC', end_color='8FAADC', fill_type='solid')
        yellowFill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')
        greenFill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        dgFill = PatternFill(start_color='52AC95', end_color='52AC95', fill_type='solid')
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        font = Font(name='Arial', b=True, color="000000")
        # Color = FFFFFF IS WHITE
        font2 = Font(name='Arial', b=True, color="FFFFFF")
        font3 = Font(name='Arial', b=False, color="000000")
        al = Alignment(wrap_text=True, horizontal="center", vertical="center")

        # Freezing Panes
        ws.freeze_panes = "C3"

        # Adjusting column widths
        ws.row_dimensions[2].height = 131
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 27
        ws.column_dimensions['C'].width = 19
        ws.column_dimensions['D'].width = 11
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12
        ws.column_dimensions['M'].width = 12
        ws.column_dimensions['N'].width = 12
        ws.column_dimensions['O'].width = 12
        ws.column_dimensions['P'].width = 10
        ws.column_dimensions['Q'].width = 10
        ws.column_dimensions['R'].width = 10
        ws.column_dimensions['S'].width = 10

        # Setting up AutoFilter
        filter_range = 'B2:S' + str(df.shape[1] + 2)
        ws.auto_filter.ref = filter_range

        # Setting Up First Two Rows
        ws['A2'] = '#'
        # Date Range: 24-March-2018 To 08-April-2018 (16 days)
        ws['A1'] = "Date Range: "+ days_range + "("+ str(no_of_days) +" days)"
        merge_and_style_range(ws, 'A1:E1', border=border, font=font, alignment=al)

        ws['F1'] = 'Actual Generation'
        merge_and_style_range(ws, 'F1:I1', border=border, fill=yellowFill, font=font2, alignment=al)

        ws['J1'] = 'Expected Generation & Irradiance'
        merge_and_style_range(ws, 'J1:O1', border=border, fill=greenFill, font=font2, alignment=al)

        ws['P1'] = 'Achievements'
        merge_and_style_range(ws, 'P1:S1', border=border, fill=dgFill, font=font2, alignment=al)

        if ws['E2'].value == 'Date':
            print "FORMATTING MERGING"
            cell_range = 'B3:B' + str(df.shape[0] + 1)
            merge_and_style_range(ws, cell_range, border=border, font=font3, alignment=al)
            cell_range = 'C3:C' + str(df.shape[0] + 1)
            merge_and_style_range(ws, cell_range, border=border, font=font3, alignment=al)
            cell_range = 'D3:D' + str(df.shape[0] + 1)
            merge_and_style_range(ws, cell_range, border=border, font=font3, alignment=al)

        for i in range(1, 6) + range(16, 20):
            ws.cell(row=2, column=i).fill = blueFill
            ws.cell(row=2, column=i).alignment = al
            ws.cell(row=2, column=i).border = border
            ws.cell(row=2, column=i).font = font

        for i in range(6, 10):
            ws.cell(row=2, column=i).fill = yellowFill
            ws.cell(row=2, column=i).alignment = al
            ws.cell(row=2, column=i).border = border
            ws.cell(row=2, column=i).font = font

        for i in range(10, 16):
            ws.cell(row=2, column=i).fill = greenFill
            ws.cell(row=2, column=i).alignment = al
            ws.cell(row=2, column=i).border = border
            ws.cell(row=2, column=i).font = font

        for i in range(1, df.shape[0] + 2):
            ws.cell(row=i, column=1).alignment = al
            ws.cell(row=i, column=1).border = border
            ws.cell(row=i, column=1).font = font

        for row in range(1, df.shape[0] + 2):
            for col in range(1, df.shape[1] + 2):
                ws.cell(row=row, column=col).alignment = al
                ws.cell(row=row, column=col).border = border
                ws.cell(row=row, column=col).font = font3
        red_color = 'ffe1e5'
        red_color_font = '9c0103'
        green_color = 'C5E0B4'

        red_font = styles.Font(color=red_color_font)
        red_fill = styles.PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
        light_green_fill = styles.PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')

        # Add a conditional formatting based on a cell comparison
        # Format if cell is less than 'formula' applied
        ws.conditional_formatting.add('P3:S' + str(df.shape[0] + 1),
                                      CellIsRule(operator='lessThan', formula=['87'], fill=red_fill, font=red_font))
        ws.conditional_formatting.add('P3:S' + str(df.shape[0] + 1),
                                      CellIsRule(operator='greaterThan', formula=['87'], fill=light_green_fill))
        return ws
    except Exception as e:
        print "Exception in ExcelNewFormat : " + str(e)
        return "ExcelFailed"


def cron_send_pvsyst_report_cleanmax(st, et, user, plant_slugs=None, recepient_email=None):
    try:
        logger.debug("hello...")
        # st = datetime.datetime(2018, 3, 24, 0, 0)
        # et = datetime.datetime(2018, 3, 28, 0, 0)
        starttime = st
        endtime = et
        no_of_days = (et - st).days + 1
        days_range = st.strftime("%d-%B-%Y") +" To "+et.strftime("%d-%B-%Y")
        # user = User.objects.get(email='dhananjay.nandedkar@cleanmaxsolar.com')
        if plant_slugs == None:
            plants = SolarPlant.objects.filter(groupClient=user.role.dg_client)
        else:
            plants = SolarPlant.objects.filter(slug__in = plant_slugs)

        file_location = "/var/tmp/monthly_report/PVsyst/"
        name = (str(user.role.dg_client) + "-" + days_range + "-PVsyst-Achievement-Analysis.xlsx").replace(" ", "-").replace(
            ",", "-")
        file_name = file_location + name

        t1 = datetime.datetime.now()
        df_dict = df_for_pvsyst_analysis(st, et, user, plants)
        t2 = datetime.datetime.now()
        print (t2 - t1)
        print "=========FINAL DF DICT=========="
        print df_dict
        flag=0
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Summary'
            df = df_dict.pop('Summary', None)
            ExcelPvsystFormat(df, ws, days_range, no_of_days)
            for key in df_dict:
                df = df_dict[key]
                ws = wb.create_sheet(key)
                ExcelPvsystFormat(df, ws, days_range, no_of_days)
            if recepient_email!=None:
                wb.save(file_name)

            print "***** Done Done Done Done Done *****"
        except Exception as e:
            print "+++++ Some Problem Some Problem +++++ %s " % e
            flag=1

        # Call email sending function to send report

        if flag==1:
            print "Some Problem found while creating Excel Cannot send email.", user.email, st, et
        else:
            print ("pvsyst achievement calculated successfully###")
            if recepient_email==None:
                return wb
            else:
                send_detailed_report_with_attachment_for_pvsyst_cleanmax(plants[0],name, recepient_email)
                print "############# Summary cron_send_pvsyst_report_cleanmax ############"
                print "Fetched Data for User ", user.email
                print "StatTime and EndTime ", st, et
                print "Data Saved to File ", name
                print "Mail is successfully Sent to : ", recepient_email


    except Exception as e:
        print "EXCEPTION IN cron_send_pvsyst_report_cleanmax"+str(e)

